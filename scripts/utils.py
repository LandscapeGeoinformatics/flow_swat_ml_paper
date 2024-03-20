# Import packages
import os
from typing import List, Tuple

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import r2_score, mean_squared_error
from joblib import dump, load
import matplotlib.pyplot as plt
import shap
import seaborn as sns
from hydroeval import evaluator, nse

sns.set_theme()


# Set project directory
def set_project_dir():
    try:
        os.chdir(r'\\export.hpc.ut.ee\gis\flow_swat_ml_paper')
    except OSError:
        os.chdir('D:/flow_swat_ml_paper')


# Get Excel file based on country code
def get_excel_file(country_code: str) -> str:
    return f'ml/{country_code}/source_data/{country_code}.xlsx'


# Add missing dates to DataFrame
def add_missing_dates(df: pd.DataFrame) -> pd.DataFrame:
    min_date = df['Date'].min().strftime('%Y-%m-%d')
    max_date = df['Date'].max().strftime('%Y-%m-%d')
    index = pd.date_range(min_date, max_date)
    df = df.set_index('Date').reindex(index, fill_value=np.nan)
    df = df.reset_index().rename(columns={'index': 'Date'})
    return df


# Get aggregation function based on time interval
def get_agg_func(variable: str, time_interval: str):
    if variable in ['Q', 'Tmax', 'Tmin']:
        return 'mean'
    elif variable == 'Pcp':
        if time_interval.lower() == 'd':
            return 'mean'
        elif time_interval.lower() == 'm':
            return 'sum'


# Read flow data from Excel file
def read_flow_data(country_code: str, time_interval: str) -> pd.DataFrame:
    excel_file = get_excel_file(country_code)
    flow_df = pd.read_excel(excel_file, sheet_name='Flow')
    flow_df = flow_df.dropna(subset=['Date']).reset_index(drop=True)
    flow_df = add_missing_dates(flow_df)
    variable = 'Q'
    agg_func = get_agg_func(variable, time_interval)
    flow_df = flow_df.set_index('Date')
    flow_df = flow_df.resample(time_interval).agg(agg_func).reset_index().round(1)
    flow_df = flow_df.rename(columns={flow_df.columns[1]: f'{variable}_{time_interval.lower()}'})
    return flow_df


# Read predictor variable data from Excel file
def read_pred_var_data(country_code: str, time_interval: str, na_values=None) -> pd.DataFrame:
    excel_file = get_excel_file(country_code)
    sheet_names = load_workbook(excel_file, read_only=True).sheetnames
    pred_var_df_list = []
    invalid_df_list = []
    variables = ['Pcp', 'Tmax', 'Tmin']
    for sheet_name in sheet_names:
        if sheet_name != 'Flow':
            df = pd.read_excel(excel_file, sheet_name=sheet_name, na_values=na_values)
            df = df.rename(
                columns={df.columns[1]: variables[0], df.columns[2]: variables[1], df.columns[3]: variables[2]}
            )
            df = df.dropna(subset=['Date']).reset_index(drop=True)
            df = add_missing_dates(df)
            df['Station'] = sheet_name
            # Get indices of dates with invalid minimum temperature values
            indices = df.loc[df['Tmax'] < df['Tmin']].index
            # Replace temperature values on those dates with NaN
            if len(indices) > 0:
                invalid_df_list.append(df.loc[indices])
                df.loc[indices, 'Tmax'] = np.nan
                df.loc[indices, 'Tmin'] = np.nan
            df = df.set_index('Date')
            # Fill missing temperature values
            if df['Tmax'].isna().sum() > 0:
                df['Tmax'] = df['Tmax'].interpolate('time')
            if df['Tmin'].isna().sum() > 0:
                df['Tmin'] = df['Tmin'].interpolate('time')
            agg_func_dict = {}
            column_name_dict = {}
            for variable in variables:
                agg_func_dict[variable] = get_agg_func(variable, time_interval)
                column_name_dict[variable] = f'{variable}_{time_interval}'
            df = df.resample(time_interval).agg(agg_func_dict).reset_index().round(1)
            df = df.rename(columns=column_name_dict)
            pred_var_df_list.append(df)
    if len(invalid_df_list) > 0:
        invalid_df = pd.concat(invalid_df_list).reset_index(drop=True)
        invalid_df.to_csv(f'ml/{country_code}/source_data/{country_code}_Tmin_invalid.csv', index=False)
    pred_var_df = pd.concat(pred_var_df_list).groupby('Date').agg('mean').reset_index().round(1)
    return pred_var_df


# Get DataFrame with lags
def calc_lags(
        df: pd.DataFrame, variables: List[str], time_interval: str, lag_width: int, n_lags: int
) -> pd.DataFrame:
    lags = df.copy()
    for variable in variables:
        for i in range(1, lag_width + n_lags, lag_width):
            lags[f'{variable}_{time_interval}-{i}'] = lags[f'{variable}_{time_interval}'].shift(i)
    return lags


# Get number of lags based on time interval
def get_n_lags(time_interval: str) -> int:
    if time_interval.lower() == 'd':
        return 28
    elif time_interval.lower() == 'm':
        return 12


# Get DataFrame with rolling aggregates
def calc_rolling_aggs(
        df: pd.DataFrame, variables: List[str], time_interval: str, window_width: int, n_windows: int,
        agg_func_list: List[str]
) -> pd.DataFrame:
    rolling_aggs = df.copy()
    for variable in variables:
        for i in range(window_width, window_width * n_windows + window_width, window_width):
            for agg_func in agg_func_list:
                agg_col = f'{variable}_{time_interval}-{i}_rolling_{agg_func}'
                rolling_aggs[agg_col] = rolling_aggs[f'{variable}_{time_interval}'].rolling(i).agg(agg_func).round(1)
    return rolling_aggs


# Get window width based on time interval
def get_window_width(time_interval: str) -> int:
    if time_interval.lower() == 'd':
        return 7
    elif time_interval.lower() == 'm':
        return 3


# Get DataFrame with leads
def calc_leads(
        df: pd.DataFrame, variables: List[str], time_interval: str, lead_width: int, n_leads: int
) -> pd.DataFrame:
    leads = df.copy()
    for variable in variables:
        for i in range(1, lead_width + n_leads, lead_width):
            leads[f'{variable}_{time_interval}+{i}'] = leads[f'{variable}_{time_interval}'].shift(-i)
    return leads


# Get path of model directory
def get_model_dir(country_code: str, target: str, feat_set: str) -> str:
    model_dir = f'ml/{country_code}/models/rf/{country_code}_{target}_{feat_set}'
    if not os.path.exists(model_dir):
        os.makedirs(model_dir)
    return model_dir


# Get catchment name based on country code
def get_catchment_name(country_code: str) -> str:
    if country_code == 'ESP':
        return 'Argos'
    elif country_code == 'EST':
        return 'Porijõgi'
    elif country_code == 'ETH':
        return 'Rib'
    elif country_code == 'USA':
        return 'Bald Eagle'


# Get target based on feature set
def get_target(feat_set: str) -> str:
    time_interval = feat_set[-1]
    if time_interval == 'd':
        return f'Q_{time_interval}+1'
    elif time_interval == 'm':
        return f'Q_{time_interval}+1'


# Get feature set list based on target
def get_feat_set(target: str) -> str:
    if target == 'Q_d+1':
        return ['FS1_d', 'FS2_d', 'FS3_d']
    elif target == 'Q_m+1':
        return ['FS1_m', 'FS2_m', 'FS3_m']


# Get model name based on feature set
def get_model_name(feat_set: str) -> str:
    time_interval = feat_set[-1].lower()
    if time_interval == 'd':
        return f'DM{feat_set[2]}'
    elif time_interval == 'm':
        return f'MM{feat_set[2]}'


# Create SHAP summary plot
def plot_shap(country_code: str, feat_set: str, test_size: float):

    # Target
    target = get_target(feat_set)

    # Model directory
    model_dir = get_model_dir(country_code, target, feat_set)

    # Test size as integer
    test_size_int = int(test_size * 100)

    # Read training and test data
    X_train = pd.read_csv(f'{model_dir}/{country_code}_{target}_{feat_set}_feat_train_{test_size_int}.csv')\
        .drop('Index', axis=1)
    X_test = pd.read_csv(f'{model_dir}/{country_code}_{target}_{feat_set}_feat_test_{test_size_int}.csv')\
        .drop('Index', axis=1)
    y_train = pd.read_csv(f'{model_dir}/{country_code}_{target}_{feat_set}_target_train_{test_size_int}.csv')
    y_test = pd.read_csv(f'{model_dir}/{country_code}_{target}_{feat_set}_target_test_{test_size_int}.csv')

    # Load model
    regressor = load(f'{model_dir}/{country_code}_{target}_{feat_set}_{test_size_int}.joblib')

    # Create DataFrame of SHAP values and export to CSV
    explainer = shap.TreeExplainer(regressor)
    shap_values = shap.TreeExplainer(regressor).shap_values(X_test)
    shap_df = pd.DataFrame(
        list(zip(X_test.columns, np.abs(shap_values).mean(0))), columns=['feature', 'abs_mean_shap']
    )
    shap_df = shap_df \
        .sort_values(by=['abs_mean_shap'], ascending=False) \
        .reset_index(drop=True)
    shap_df.to_csv(f'{model_dir}/{country_code}_{target}_{feat_set}_shap_{test_size_int}.csv', index=False)

    # Get indices of top 10 features
    feature_names = shap_df.head(10)['feature'].to_list()
    indices = []
    for feat in feature_names:
        index = X_test.columns.to_list().index(feat)
        indices.append(index)

    # Read results
    results_df = pd.read_csv(f'ml/{target}_rf_metrics.csv')

    # Get NSE calculated on test data
    subset = results_df[(results_df['country_code'] == country_code) & (results_df['feat_set'] == feat_set)]\
        .reset_index(drop=True)
    nse_test = float(subset[subset['test_size_int'] == test_size_int]['nse_test'].values[0])

    # SHAP summary plot
    shap.summary_plot(shap_values[:, indices], X_test.iloc[:, indices], feature_names=feature_names, show=False)
    fig = plt.gcf()
    title_label_top = get_catchment_name(country_code)
    title_label_bottom = (
        'Target: ' + f"$\mathregular{{{target.split('_')[0]}}}_\mathregular{{{target.split('_')[1]}}}$" +
        '$\quad$' +
        'Model: ' + get_model_name(feat_set) +
        '$\quad$' +
        f'Test size: {test_size_int}%' +
        '$\quad$' +
        'NSE=' + f'{nse_test}'
    )
    plt.title(title_label_top + '\n' + title_label_bottom)
    plt.tight_layout()
    plt.savefig(
        f'ml/figures/shap/{country_code}_{target}_{feat_set}_shap_summary_{test_size_int}.png', dpi=300,
        bbox_inches='tight'
    )
    plt.close(fig)

    return


# Concatenate RF metrics
def concat_rf_metrics(time_interval: str) -> pd.DataFrame:

    set_project_dir()

    target = f'Q_{time_interval}+1'
    feat_set_list = [f'FS{i}_{time_interval}' for i in range(1, 4)]
    country_codes = ['ESP', 'USA', 'EST', 'ETH']

    df_list = []
    for feat_set in feat_set_list:
        for country_code in country_codes:
            df = pd.read_csv(
                f'ml/{country_code}/models/rf/{country_code}_{target}_{feat_set}/'
                f'{country_code}_{target}_{feat_set}_results.csv'
            )
            df_list.append(df)
    results_df = pd.concat(df_list).reset_index(drop=True)

    results_df['training_time_seconds'] = pd.to_timedelta(results_df['training_time']).dt.total_seconds()

    catchment_names = [get_catchment_name(country_code) for country_code in country_codes]
    catchment_name_dict = dict(zip(country_codes, catchment_names))
    results_df['catchment_name'] = results_df['country_code'].replace(catchment_name_dict)
    results_df = results_df.sort_values(['catchment_name', 'feat_set']).reset_index(drop=True)

    return results_df


# Get SWAT nRMSE based on country code and time interval
def get_swat_nrmse(country_code: str, time_interval: str) -> Tuple[float, float]:
    set_project_dir()
    df = pd.read_excel('swat/Nrmse.xlsx')
    keep_indices = [0]
    country_index = list(df.columns).index(country_code)
    keep_indices.extend([country_index, country_index + 1])
    drop_indices = [i for i in range(len(df.columns)) if i not in keep_indices]
    df = df.drop(df.columns[drop_indices], axis=1)
    df = df.rename(columns=df.iloc[0])
    df = df.drop(df.index[0]).reset_index(drop=True)
    nrmse_cal = df[time_interval.upper()].iloc[0]
    nrmse_val = df[time_interval.upper()].iloc[1]
    return nrmse_cal, nrmse_val


# Get SWAT NSE based on country code and time interval
def get_swat_nse(country_code: str, time_interval: str) -> Tuple[float, float]:
    set_project_dir()
    col_names = ['Model', 'NSE_Cal', 'RMSE_Cal', 'NSE_Val', 'RMSE_Val']
    df = pd.read_excel('swat/SWAT_metrics.xlsx', names=col_names, skiprows=1)
    nse_cal = round(df[df['Model'] == f'{country_code}_{time_interval.upper()}']['NSE_Cal'].values[0], 2)
    nse_val = round(df[df['Model'] == f'{country_code}_{time_interval.upper()}']['NSE_Val'].values[0], 2)
    return nse_cal, nse_val


# Get start and end date of training period
def get_train_period(country_code: str):
    if country_code == 'ESP':
        return '2006-01-01', '2010-12-31'
    elif country_code == 'USA':
        return '2002-01-01', '2006-12-31'
    elif country_code == 'EST':
        return '2009-01-01', '2013-12-31'
    elif country_code == 'ETH':
        return '1997-01-01', '2001-12-31'


# Get start and end date of testing period
def get_test_period(country_code: str):
    if country_code == 'ESP':
        return '2012-01-01', '2016-12-31'
    elif country_code == 'USA':
        return '2008-01-01', '2012-12-31'
    elif country_code == 'EST':
        return '2015-01-01', '2019-12-31'
    elif country_code == 'ETH':
        return '2003-01-01', '2007-12-31'
